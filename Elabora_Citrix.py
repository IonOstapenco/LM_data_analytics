 # Python Classi dati server Citrix
"""
Această procedură prelucrează fișiere Citrix (.xlsx) și extrage:
•	lista de servere Citrix
•	lista de VM-uri Citrix
•	le filtrează
•	creează 2 fișiere CSV curate

"""
#from sys import exit
import openpyxl as xl
import csv
import Common as cm

# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
class c_Server:
    def __init__(self, nome, attributo):
        self.nome = nome
        self.dati = attributo

Server = []
# campuri Server
Server_field = [
             "name",         # colonna A 1         
             "address",      # address colonna K 11      
             "cpus",         # CPUS colonna O 15  
             "description",  # descriptions AO 41  
             "funzione",     # Riepilogo pool       
             "pool"          # Riepilogo pool            
            ]

class c_Vms:
    def __init__(self, nome, attributo):
        self.nome = nome
        self.dati = attributo

# -- 
Vms = []
# campuri masini virtuale
Vms_field = [
            "name",           # colonna A 1               
             "address",       # colonna G 7             
             "cpus",          # colonna AF 32               
             "funzione",      # Riepilogo pool                           
             "operating_system",   # colonna U 21  
             "pool",          # Riepilogo pool               
             "power_state",   # colonna B 2        
             "running_on"     # colonna D 4                          
            ]

# colectia/lista pentru excludere 
ExcludeSRV = ["passivo produzione",
           "sviluppo",
           "dr - bolla - wpr"
           ] # ---> mediile pentru excludere 
Exclude = ["Halted"] # ---> tipa lista de masini oprite 
# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=

cm.check_outdir(cm.out_path)
# ===================================================================
# 3️⃣ SELECTAREA FIȘIERULUI
# =======================================================================
cm.list_files_scandir(cm.start_path, cm.pr["Citrix_files_Pattern"], cm.pr['XLS_end'])

mesg = "Elemento [{}]"
mes1 = "FILE:[{}]"

temp = sorted(cm.files, reverse=True)
print("Elaborazione file " + temp[0])


# =============================================================================
# 4️⃣ CITIRE EXCEL (openpyxl) 
# ============================================================================
wb = xl.load_workbook(temp[0]) # --> citim fisierul excel

wbs = wb.sheetnames #  --> denumiri la Sheets 
n = len(wbs)  

#================================================================================
# 5️⃣ CONSTRUIRE MAPARE POOL 
# ========================================================================
ws = wb['Riepilogo'] # --> se lucreaza pe sheeytul Riepilogo 
pool = {# --> !! prea complicat, dupa pozitie--
    # de dorit sa se citeasca DUPA NUME!!! -- poate ca aici https://stackoverflow.com/questions/58398464/read-excel-sheet-table-listobject-into-python-with-pandas
    str(ws.cell(row=6, column=2).value).lower(): str(ws.cell(row=6, column=3).value).lower(), # --> se ia cheia B6 si valoarea C6 (exemplu: XE01CP1x, PRODUZIONE)
    str(ws.cell(row=7, column=2).value).lower(): str(ws.cell(row=7, column=3).value).lower(),
    str(ws.cell(row=8, column=2).value).lower(): str(ws.cell(row=8, column=3).value).lower(),
    str(ws.cell(row=9, column=2).value).lower(): str(ws.cell(row=9, column=3).value).lower(),
    str(ws.cell(row=6, column=4).value).lower(): str(ws.cell(row=6, column=5).value).lower(),
    str(ws.cell(row=7, column=4).value).lower(): str(ws.cell(row=7, column=5).value).lower(),
    str(ws.cell(row=8, column=4).value).lower(): str(ws.cell(row=8, column=5).value).lower(),
    str(ws.cell(row=9, column=4).value).lower(): str(ws.cell(row=9, column=5).value).lower(),
    str(ws.cell(row=10, column=4).value).lower(): str(ws.cell(row=10, column=5).value).lower()
}

for w in range(1,n):
    ws = wb[wbs[w]]

    for i in range(1,ws.max_row):
        if str(ws.cell(row=i, column=1).value).lower() == "servers":
            i += 2
            while ( str(ws.cell(row=i, column=1).value).lower() != 'none' ):
                print('\r', mesg.format(i), end='', flush=True)
                nome = str(ws.cell(row=i, column=1).value).lower()
                l = []
                l.append(nome)
                l.append(ws.cell(row=i, column=11).value)
                if ( len(str(ws.cell(row=i, column=15).value)) < 4 ): 
                    l.append('n.d.')
                    l.append('-')
                else:
                    a = str(ws.cell(row=i, column=15).value).lower().split(" ")
                    l.append(a[2])
                    l.append(ws.cell(row=i, column=41).value)

                a = str(wbs[w]).lower()
                l.append(pool[a])
                l.append(a)

                z = c_Server(nome,l)
                Server.append(z)
                i += 1
            #  FINE Loop Elenco SERVERS

        if str(ws.cell(row=i, column=1).value).lower() == "vms":
            i += 2
            while ( str(ws.cell(row=i, column=1).value).lower() != 'none' ):
                print('\r', mesg.format(i), end='', flush=True)
                nome = str(ws.cell(row=i, column=1).value).lower()
                l = []
                l.append(nome)
                l.append(ws.cell(row=i, column=7).value)
                if ( len(str(ws.cell(row=i, column=32).value)) < 4 ): 
                    l.append('n.d.')
                
                else:
                    a = str(ws.cell(row=i, column=32).value).lower().split(" ")
                    l.append(a[2])
                a = str(wbs[w]).lower()
                l.append(pool[a])
                l.append(ws.cell(row=i, column=21).value)
                l.append(a)
                l.append(ws.cell(row=i, column=2).value)
                l.append(ws.cell(row=i, column=4).value)
                z = c_Vms(nome,l)
                Vms.append(z)
                i += 1
#               #  FINE Loop Elenco VMs
            break
#   i += 1   Così siamo alla riga successiva alla riga vuota trovata dopo il blocco "servers"
#
# - Riprendiamo a scorrere la prima colonna e cerchiamo una cella con il valore "VMs" 
# - Da qui in poi c'è l'elenco dei server virtuali.




# -=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=


#
# Scrittura file intermedio: Server_list.csv
#

y = [cm.out_path, cm.pr["OUT_Server_Citrix"]]

output_file = cm.dr.join(y)

f = open(output_file, "w")

f.write("sep=" + cm.cs + "\n")

print("\n\nScrittura file " + output_file)

riga = cm.cs.join(Server_field)

f.write(riga + "\n")

n = len(Server)
for j in range(0, n):
    print('\r', mesg.format(j), end='', flush=True)
    riga = cm.cs.join(map(str, Server[j].dati))

    if Server[j].dati[4] not in ExcludeSRV:
        f.write(riga + "\n")
f.close()
print ("\n>-----------------------------------------------------------<\n")


#
# Scrittura file intermedio: Vms_list.csv
#

y = [cm.out_path, cm.pr["OUT_Vms_Citrix"]]

output_file = cm.dr.join(y)

f = open(output_file, "w")

f.write("sep=" + cm.cs + "\n")

print("\n\nScrittura file " + output_file)

riga = cm.cs.join(Vms_field)

f.write(riga + "\n")

n = len(Vms)
for j in range(0, n):
    print('\r', mesg.format(j), end='', flush=True)
    riga = cm.cs.join(map(str, Vms[j].dati))

    if Vms[j].dati[3].lower() not in ExcludeSRV:
        if Vms[j].dati[6] not in Exclude:
            f.write(riga + "\n")
f.close()
print ("\n>-----------------------------------------------------------<\n")