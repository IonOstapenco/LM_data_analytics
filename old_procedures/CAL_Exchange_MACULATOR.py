import os
import csv
import json
import pandas as pd
from openpyxl import Workbook
from pathlib import Path

# ---------------------
# Citim parametri.json
with open("parametri.json", encoding="utf-8") as f:
    config = json.load(f)

SOURCE_DIR = config["Source_dir"]
REPORT_DIR = config["Report_dir"]

#am adaugat sa se salveze in output
OUTPUT_PATH = Path(SOURCE_DIR) / REPORT_DIR / config["Output_dir"]

#AD_SUBFOLDER = r"Active Directory Results - SERVIZICED - 2.12.2025-151753\ServiziCed.local"  # momentan hardcodat aici, hz cum sa pun in parametri.json

#added in Parametri,json
#adaugate in Parametri.json
AD_BASE_PATTERN = config["AD_Base_Pattern"]
AD_DOMAIN_SUBFOLDER = config["AD_Domain_Subfolder"]

# ---------------------
#  detectăm automat cel mai recent folder AD
# detectam automat cel mai "proaspat"/recent folder AD
report_root = Path(SOURCE_DIR) / REPORT_DIR

ad_folders = [
    p for p in report_root.iterdir()
    if p.is_dir() and p.name.startswith(AD_BASE_PATTERN)
]

#se non lo e, visualiziamo il mesaggio
# daca nu este atunci afisam mesajul
if not ad_folders:
    raise FileNotFoundError("! Atenzione! Non c'E nessuna cartella con quest nome. (Nu este niciun folder) Active Directory Results")

latest_ad_folder = max(ad_folders, key=lambda p: p.stat().st_mtime)

#trecem pe submape- ca in exercitiu
root_folder = latest_ad_folder / AD_DOMAIN_SUBFOLDER

print(f"[INFO] AD cartella scelta automaticamente/  AD folder ales automat: {root_folder}")

# ---------------------
# Output files
# Ho cambiato il nome in modo che la prossima volta inizi a elaborare i file che iniziano con AD
# am schimbat denumirea ca data viitoare incepe sa se prelucreze fisiere care se incep cu AD
#output_file = os.path.join(SOURCE_DIR, REPORT_DIR, "usersAndGroupsResultFinale.csv")
#output_file_xlsx = os.path.join(SOURCE_DIR, REPORT_DIR, "usersAndGroupsResultFinale.xlsx")


#ca sa se salveze in output
#output_file = OUTPUT_PATH / "usersAndGroupsResultFinale.csv"
output_file_xlsx = OUTPUT_PATH / "usersAndGroupsResultFinale.xlsx"


input_cn_file = os.path.join(
    SOURCE_DIR,
    REPORT_DIR,
    "output-1",
    config["AD_Users_Results_attivi"]
)

keywords_exclude = ["CR", "AX", "FI", "FJ", "test", "Administrator", "Cedacri_Admin", "harvest1", "PSMConnectCED", "R-xx1", "R-XZNETW"]


#normalizam keywords de excludere 
keywords_exclude_norm = [k.lower() for k in keywords_exclude]


# ---------------------
#legere files AD-usersAndGroupsResult.csv
# citim fisiere AD-usersAndGroupsResult.csv
users = set()  # (UserName, SamAccountName)

for root, dirs, files in os.walk(root_folder):
    for file in files:
        if file == "AD-usersAndGroupsResult.csv":
            file_path = os.path.join(root, file)
            try:
                ##c'è stato un errore con utf-8, ho provato con gli altri e con utf-16 ha funzionato
                #era eroare cu utf-8, am incercat cu celelalte si cu utf-16 a m,ers
                with open(file_path, newline="", encoding="utf-16") as csvfile:
                    reader = csv.DictReader(csvfile, delimiter=";")

                    if (
                        "UserName" not in reader.fieldnames
                        or "SamAccountName" not in reader.fieldnames
                    ):
                        print(f"[ATTENZIONE] Coloana lipseste in {file_path}")
                        continue

                    for row in reader:
                        username = row["UserName"].strip()
                        sam = row["SamAccountName"].strip()

                        # excluderi SamAccountName
                         # excluderi SamAccountName
                        if sam and any(sam.upper().startswith(p) for p in keywords_exclude):
                            continue

                        if username or sam:
                            users.add((username, sam))

            except Exception as e:
                print(f"[ERROR] {file_path} -> {e}")

# ---------------------
#scrivere CSV
# scriem CSV
#with open(output_file, "w", newline="", encoding="utf-8") as csvfile:
#    writer = csv.writer(csvfile)
#    writer.writerow(["UserName", "SamAccountName"])
#    for username, sam in sorted(users):
#        writer.writerow([username, sam])

#print(f"Fisier CSV generat: {output_file}")
#print(f"Total UserName unicale: {len(users)}")

# ---------------------
#scrivere Excel
# scriem Excel
wb = Workbook()

#main Sheet -- il principale sheet
# sheet principal
ws_main = wb.active
ws_main.title = "AD-usersAndGroupsResultFinale"
ws_main.append(["UserName", "SamAccountName"])

for username, sam in sorted(users):
    ws_main.append([username, sam])

# autosize
for col in ws_main.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws_main.column_dimensions[col[0].column_letter].width = max_length + 2

# ---------------------
# sheet Results
ws_results = wb.create_sheet(title="Results")
ws_results.append(["CN_00", "UserName", "InnerJoin"])

# ---------------------
#legiamo CN_00
# citim CN_00
cn_list = []
try:
    cn_df = pd.read_csv(
        input_cn_file,
        sep="|",
        encoding="utf-8",
        skiprows=1,
        dtype=str
    ).fillna("")

    if "CN_00" not in cn_df.columns:
        print("[ERROR] Coloana CN_00 lipseste")
        cn_df = pd.DataFrame(columns=["CN_00"])
    else:
        cn_list = [cn.strip() for cn in cn_df["CN_00"] if cn.strip()]
        cn_df = pd.DataFrame({"CN_00": cn_list})

except Exception as e:
    print(f"[ERROR] Citire CN_00 -> {e}")
    cn_df = pd.DataFrame(columns=["CN_00"])

# ---------------------
# Users dataframe
username_list = [u.strip() for u, _ in sorted(users) if u.strip()]
users_df = pd.DataFrame({"UserName": username_list})

# normalizare
cn_df["CN_00_norm"] = cn_df["CN_00"].str.lower().str.strip()
users_df["UserName_norm"] = users_df["UserName"].str.lower().str.strip()

#mai aplicam si exlcudere in UserName_norm -- cu expresia lambda
users_df = users_df[
    ~users_df["UserName_norm"].apply(
        lambda x: any(k in x for k in keywords_exclude_norm)
    )
]


# *-*-*-*-*-*
# INNER JOIN (case-insensitive)

merged_df = pd.merge(
    cn_df,
    users_df,
    left_on="CN_00_norm",
    right_on="UserName_norm",
    how="inner"
)


#cream o lista Users cu valori din InnerJoion






'''
#cream inner join ordonat
# doar INNER JOIN + sortare alfabetica
merged_df_sorted = (
    merged_df
    .loc[:, ["CN_00", "UserName"]]
    .dropna()
    .drop_duplicates()
    .sort_values(by="UserName", key=lambda col: col.str.lower())
)
'''



print(f"Numar recorduri INNER JOIN: {len(merged_df)}")
print(f"Numero di record INNER JOIN: {len(merged_df)}")

# ---------------------
# populam Results
max_len = max(len(cn_list), len(username_list))

for i in range(max_len):
    cn_val = cn_list[i] if i < len(cn_list) else ""
    username_val = username_list[i] if i < len(username_list) else ""

    inner_val = (
        cn_val
        if cn_val.lower() in merged_df["CN_00_norm"].values
        else ""
    )


    ws_results.append([cn_val, username_val, inner_val])



    

# autosize Results
for col in ws_results.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws_results.column_dimensions[col[0].column_letter].width = max_length + 2


# ---------------------
# sheet Results2 - DOAR INNER JOIN pe UserName
ws_results2 = wb.create_sheet(title="Results2")
ws_results2.append(["USERS"])

# cream lista de UserName din INNER JOIN, sortata alfabetic
users_innerjoin_sorted = merged_df["UserName"].dropna().drop_duplicates().sort_values(key=lambda col: col.str.lower())

for user in users_innerjoin_sorted:
    ws_results2.append([user])

# autosize Results2
for col in ws_results2.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws_results2.column_dimensions[col[0].column_letter].width = max_length + 2




''' 
# ---------------------
# sheet Results2 - DOAR INNER JOIN
ws_results2 = wb.create_sheet(title="Results2")
ws_results2.append(["CN_00", "USERS"])

for _, row in merged_df_sorted.iterrows():
    ws_results2.append([row["CN_00"], row["UserName"]])

# autosize Results2
for col in ws_results2.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws_results2.column_dimensions[col[0].column_letter].width = max_length + 2   


'''    

# ---------------------
#salvataggio Excel
# salvare Excel
wb.save(output_file_xlsx)
print(f"Fisier Excel generat: {output_file_xlsx}")
